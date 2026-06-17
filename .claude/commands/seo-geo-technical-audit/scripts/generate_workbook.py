"""
generate_workbook.py  —  seo-geo-technical-audit-v9
=====================================================
Reads workbook_data.json (produced by extract_data.py) and generates
the audit workbook .xlsx with up to 4 tabs:
  Tab 1 — Overview
  Tab 2 — All Phases
  Tab 3 — Schema Analysis (optional — set INCLUDE_SCHEMA = True)
  Tab 4 — Action Plan

HOW TO USE
----------
1. Run extract_data.py first to produce workbook_data.json.
2. Set AUDIT_DIR below to match the client project root.
3. Set OUTPUTS_SUBDIR to the subfolder where the .xlsx should be saved.
4. Set INCLUDE_SCHEMA = True if the schema CSV exists (see SCHEMA_CSV_NAME).
5. Set OUT_FILENAME_SUFFIX if you want to version the output (e.g. " v2").
6. Run: python3 generate_workbook.py

CLIENT_NAME and AUDIT_DATE are read automatically from workbook_data.json.
You do not need to set them here — set them once in extract_data.py.

REQUIREMENTS
------------
  pip install openpyxl
"""

import csv
import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CLIENT CONFIGURATION ─────────────────────────────────────────────────────
# Only these variables need to be set per client. CLIENT_NAME and AUDIT_DATE
# are read from workbook_data.json — set them in extract_data.py.
AUDIT_DIR          = ""   # e.g. "/Users/username/claude_code/client_audit"
OUT_FILENAME_SUFFIX = ""  # e.g. " v2" to preserve a previous version

# Set True only if the schema CSV exists at:
# {AUDIT_DIR}/Outputs/CSV/{CLIENT_NAME} - Schema Markup Corrections - {AUDIT_DATE}.csv
INCLUDE_SCHEMA = False

IN_JSON = os.path.join(AUDIT_DIR, "Workbook", "workbook_data.json")

# ── COLOURS ───────────────────────────────────────────────────────────────────
PRIORITY_FILLS = {
    "HIGH":                         PatternFill("solid", fgColor="F4CCCC"),
    "MEDIUM":                       PatternFill("solid", fgColor="FCE8B2"),
    "LOW":                          PatternFill("solid", fgColor="B7E1CD"),
    "MONITOR":                      PatternFill("solid", fgColor="BFBFBF"),
    "Manual Verification Required": PatternFill("solid", fgColor="C9DAF8"),
}
HEADER_FILL  = PatternFill("solid", fgColor="274E13")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
PHASE_FILL   = PatternFill("solid", fgColor="D9EAD3")
RAG_RED      = PatternFill("solid", fgColor="F4CCCC")
RAG_AMBER    = PatternFill("solid", fgColor="FCE8B2")
RAG_GREEN    = PatternFill("solid", fgColor="B7E1CD")

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP   = Alignment(wrap_text=True, vertical="top")
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
BOLD   = Font(bold=True)

# ── COLUMN SPEC ───────────────────────────────────────────────────────────────
# 17 columns — v9 workbook spec. Column order matches OUTPUTS-WORKBOOK.md.
COLUMNS = [
    "Status", "Family", "Category", "Sub-Category", "Analyzed Element",
    "Description", "Score", "Weight/Importance (Numeric)", "Importance Tier",
    "Priority Score", "Priority", "Who's in charge?",
    "Score Explanation", "Data Analyzed", "How to correct?", "Comments", "Sources",
]
# Column widths (characters) — one value per column above.
COL_WIDTHS = [10, 12, 18, 22, 28, 40, 10, 10, 12, 10, 10, 16, 50, 30, 50, 40, 30]

# Canonical phase order for sorting and scorecard display.
# Must match the Family values produced by extract_data.py after normalisation.
PHASES = [
    "Technical",
    "UX",
    "Tools & Configuration",
    "On-Site SEO",
    "Off-Site",
    "GEO / AI Visibility",
]

# ── HELPERS ───────────────────────────────────────────────────────────────────
def rag_fill(pct):
    if pct < 50:  return RAG_RED
    if pct < 75:  return RAG_AMBER
    return RAG_GREEN

def set_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header(ws, row=1):
    for ci, name in enumerate(COLUMNS, 1):
        c = ws.cell(row=row, column=ci, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTRE
        c.border = BORDER
    ws.row_dimensions[row].height = 30

def apply_row_format(ws, row_idx, priority, n_cols):
    fill = PRIORITY_FILLS.get(priority)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = WRAP
        cell.border = BORDER
        if fill:
            cell.fill = fill

def row_to_values(r):
    """Map JSON keys → workbook column order (17 columns, v7 spec)."""
    return [
        r["status"], r["family"], r["category"], r["sub_category"],
        r["element"], r["description"], r["score"],
        r["weight"], r["tier"], r["priority_score"],
        r["priority"], r["who"], r["explanation"],
        r.get("data_analyzed", ""), r["how_to"], r["comments"], r["sources"],
    ]

def priority_sort_key(r):
    ps = r["priority_score"]
    return ps if isinstance(ps, (int, float)) else 999

# ── TAB 3 (OPTIONAL): SCHEMA ANALYSIS ────────────────────────────────────────
SCHEMA_COLUMNS = [
    "Page Type", "Schema Markup Type", "Implementation Status",
    "Current Schema Markup", "Recommended Schema Markup",
    "Recommendations", "Sources",
]
SCHEMA_COL_WIDTHS = [20, 28, 22, 60, 60, 50, 30]

def build_schema_tab(wb, schema_csv):
    """
    Builds the Schema Analysis tab from the schema markup corrections CSV.
    Called only when INCLUDE_SCHEMA = True.
    """
    ws = wb.create_sheet("Schema Analysis")
    for ci, name in enumerate(SCHEMA_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = BOLD
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[1].height = 20

    with open(schema_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for ri, row in enumerate(reader, 2):
            for ci, col in enumerate(SCHEMA_COLUMNS, 1):
                value = row.get(col, "")
                if col == "Recommendations":
                    value = re.sub(r'\.\s+(?=[A-Z])', '.\n', value.strip())
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.alignment = WRAP
                cell.border = BORDER

    for i, w in enumerate(SCHEMA_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

# ── TAB 2: ALL PHASES ─────────────────────────────────────────────────────────
def build_all_phases(wb, rows):
    ws = wb.create_sheet("All Phases")
    write_header(ws)
    # Sort: phase order first, then category, then sub-category within each phase.
    sorted_rows = sorted(rows, key=lambda r: (
        PHASES.index(r["family"]) if r["family"] in PHASES else 99,
        r["category"] or '',
        r["sub_category"] or ''
    ))
    for i, r in enumerate(sorted_rows, 2):
        for ci, val in enumerate(row_to_values(r), 1):
            ws.cell(row=i, column=ci, value=val)
        apply_row_format(ws, i, r["priority"], len(COLUMNS))
    set_col_widths(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

# ── TAB 4: ACTION PLAN ────────────────────────────────────────────────────────
def build_action_plan(wb, rows):
    ws = wb.create_sheet("Action Plan")
    write_header(ws)
    high_rows = sorted(
        [r for r in rows if r["priority"] == "HIGH"],
        key=priority_sort_key
    )
    for i, r in enumerate(high_rows, 2):
        for ci, val in enumerate(row_to_values(r), 1):
            ws.cell(row=i, column=ci, value=val)
        apply_row_format(ws, i, r["priority"], len(COLUMNS))
    set_col_widths(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

# ── TAB 1: OVERVIEW ───────────────────────────────────────────────────────────
def build_overview(wb, rows, phase_scores, exec_summary, client_name, audit_date):
    ws = wb.create_sheet("Overview")

    # PERFORMANCES header
    # NOTE: cell merges below (B2:M2, B8:M8, etc.) assume exactly 6 phases fit in columns B–M.
    # If phases are ever added or removed, update these ranges and the RAG key column offset.
    ws.merge_cells("B2:M2")
    c = ws["B2"]
    c.value = "PERFORMANCES"
    c.font  = Font(bold=True, size=14)
    c.alignment = CENTRE

    # Phase scorecard
    for col_offset, phase in enumerate(PHASES):
        col = col_offset + 2
        hc = ws.cell(row=4, column=col, value=phase)
        hc.font = BOLD; hc.alignment = CENTRE; hc.fill = PHASE_FILL
        pct = phase_scores.get(phase, 0)
        sc = ws.cell(row=5, column=col, value=f"{pct}%")
        sc.alignment = CENTRE; sc.font = Font(bold=True, size=12)
        sc.fill = rag_fill(pct)

    # Overall
    oc_col = len(PHASES) + 2
    ws.cell(row=4, column=oc_col, value="Overall").font = BOLD
    ws.cell(row=4, column=oc_col).alignment = CENTRE
    ws.cell(row=4, column=oc_col).fill = PHASE_FILL
    overall_pct = phase_scores.get("Overall", 0)
    oc = ws.cell(row=5, column=oc_col, value=f"{overall_pct}%")
    oc.alignment = CENTRE; oc.font = Font(bold=True, size=12)
    oc.fill = rag_fill(overall_pct)

    # RAG key
    key_col = oc_col + 2
    ws.cell(row=4, column=key_col, value="Key").font = BOLD
    for offset, (label, fill) in enumerate([
        ("To optimize (<50%)",       RAG_RED),
        ("May be optimised (<75%)",  RAG_AMBER),
        ("Optimized (>75%)",         RAG_GREEN),
    ]):
        kc = ws.cell(row=4 + offset, column=key_col, value=label)
        kc.fill = fill; kc.alignment = WRAP

    # HIGH priority items per phase
    ws.merge_cells("B8:M8")
    hh = ws["B8"]
    hh.value = "HIGH PRIORITY ITEMS BY PHASE"
    hh.font  = Font(bold=True, size=11)
    hh.fill  = PRIORITY_FILLS["HIGH"]
    hh.alignment = CENTRE

    for col_offset, phase in enumerate(PHASES):
        col = col_offset + 2
        ph = ws.cell(row=9, column=col, value=phase)
        ph.font = BOLD; ph.fill = PHASE_FILL; ph.alignment = CENTRE

        high_items = sorted(
            [r for r in rows if r["family"] == phase and r["priority"] == "HIGH"],
            key=priority_sort_key
        )
        for item_offset, item in enumerate(high_items):
            cell = ws.cell(
                row=10 + item_offset, column=col,
                value=f"• {item['element']} ({item['priority_score']})"
            )
            cell.fill = PRIORITY_FILLS["HIGH"]
            cell.alignment = WRAP

    # Executive summary block — placed below the tallest HIGH-items column.
    # 10 = first item row; +2 = one blank row gap.
    max_high = max(
        (len([r for r in rows if r["family"] == phase and r["priority"] == "HIGH"])
         for phase in PHASES),
        default=0
    )
    sum_row = max(22, 10 + max_high + 2)
    ws.merge_cells(f"B{sum_row}:M{sum_row}")
    sh = ws[f"B{sum_row}"]
    sh.value = "EXECUTIVE SUMMARY"; sh.font = Font(bold=True, size=11)
    sh.fill = PHASE_FILL; sh.alignment = CENTRE

    ws.merge_cells(f"B{sum_row+1}:M{sum_row+6}")
    sc = ws[f"B{sum_row+1}"]
    sc.value = exec_summary; sc.alignment = WRAP
    ws.row_dimensions[sum_row + 1].height = 120

    # Score distribution table
    dist_row = sum_row + 9
    ws.cell(row=dist_row, column=1, value="Score Distribution").font = BOLD
    dist_headers = [
        "Phase", "Not Present", "Weak", "Medium",
        "High", "Excellent", "Not Applicable", "Data Missing"
    ]
    for ci, h in enumerate(dist_headers, 1):
        hc = ws.cell(row=dist_row, column=ci, value=h)
        hc.font = BOLD; hc.fill = PHASE_FILL; hc.alignment = CENTRE

    score_labels = ["Not Present", "Weak", "Medium", "High", "Excellent",
                    "Not Applicable", "Data Missing"]
    for row_off, phase in enumerate(PHASES + ["All Phases"]):
        phase_rows = [r for r in rows if r["family"] == phase] \
                     if phase != "All Phases" else rows
        dr = dist_row + 1 + row_off
        ws.cell(row=dr, column=1, value=phase).font = BOLD
        for ci, label in enumerate(score_labels, 2):
            ws.cell(row=dr, column=ci,
                    value=sum(1 for r in phase_rows if r["score"] == label)
                    ).alignment = CENTRE

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 20

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    with open(IN_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    client_name  = data["client"]
    audit_date   = data["audit_date"]
    rows         = data["rows"]
    phase_scores = data["phase_scores"]
    exec_summary = data.get("exec_summary", "")

    schema_csv = os.path.join(
        AUDIT_DIR, "Outputs", "CSV",
        f"{client_name} - Schema Markup Corrections - {audit_date}.csv"
    )
    out_xls = os.path.join(
        AUDIT_DIR, "Workbook",
        f"{client_name} - SEO GEO Audit - {audit_date}{OUT_FILENAME_SUFFIX}.xlsx"
    )

    wb = Workbook()
    wb.remove(wb.active)

    build_overview(wb, rows, phase_scores, exec_summary, client_name, audit_date)
    build_all_phases(wb, rows)
    if INCLUDE_SCHEMA:
        build_schema_tab(wb, schema_csv)
    build_action_plan(wb, rows)

    os.makedirs(os.path.dirname(out_xls), exist_ok=True)
    wb.save(out_xls)

    high_count = sum(1 for r in rows if r["priority"] == "HIGH")
    print(f"Saved: {out_xls}")
    print(f"  Overview:     scorecard + HIGH items + exec summary + score distribution")
    print(f"  All Phases:   {len(rows)} rows, sorted phase → category → sub-category")
    if INCLUDE_SCHEMA:
        print(f"  Schema:       from {os.path.basename(schema_csv)}")
    print(f"  Action Plan:  {high_count} HIGH priority rows, sorted Priority Score ascending")

if __name__ == "__main__":
    main()
