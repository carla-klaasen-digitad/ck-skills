"""
build_schema_tab.py  —  seo-geo-technical-audit-v7
====================================================
Reads the schema markup corrections CSV and generates a standalone
Schema Analysis .xlsx file.

Use this script when the schema sub-phase was completed after the main
workbook was already generated. The original workbook is not modified.

If the schema sub-phase is completed before workbook generation, set
INCLUDE_SCHEMA = True in generate_workbook.py instead — the schema tab
will be included automatically as Tab 3 in the main workbook.

HOW TO USE FOR A NEW CLIENT
----------------------------
1. Set CLIENT_NAME, AUDIT_DATE, AUDIT_DIR below.
2. Run: python3 build_schema_tab.py

The schema corrections CSV must already exist at:
  {AUDIT_DIR}/Outputs/CSV/{CLIENT_NAME} - Schema Markup Corrections - {AUDIT_DATE}.csv

REQUIREMENTS
------------
  pip install openpyxl
"""

import csv
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CLIENT CONFIGURATION ─────────────────────────────────────────────────────
AUDIT_DIR   = ""  # e.g. "/Users/username/claude_code/client_audit"
CLIENT_NAME = ""  # e.g. "YoCrunch"
AUDIT_DATE  = ""  # e.g. "March 2026"

IN_CSV  = os.path.join(
    AUDIT_DIR, "Outputs", "CSV",
    f"{CLIENT_NAME} - Schema Markup Corrections - {AUDIT_DATE}.csv"
)
OUT_XLS = os.path.join(
    AUDIT_DIR, "Workbook",
    f"{CLIENT_NAME} - Schema Analysis - {AUDIT_DATE}.xlsx"
)

# ── COLUMN SPEC ───────────────────────────────────────────────────────────────
COLUMNS = [
    "Page Type",
    "Schema Markup Type",
    "Implementation Status",
    "Current Schema Markup",
    "Recommended Schema Markup",
    "Recommendations",
    "Sources",
]
COL_WIDTHS = [20, 28, 22, 60, 60, 50, 30]

# ── STYLES ────────────────────────────────────────────────────────────────────
thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP   = Alignment(wrap_text=True, vertical="top")
BOLD   = Font(bold=True)

# ── HELPERS ───────────────────────────────────────────────────────────────────
def format_recommendations(text):
    """
    Split recommendation prose into one sentence per line.
    Splits on a period followed by whitespace and an uppercase letter,
    preserving abbreviations and mid-sentence capitals.
    """
    if not text:
        return text
    return re.sub(r'\.\s+(?=[A-Z])', '.\n', text.strip())

# ── BUILD ─────────────────────────────────────────────────────────────────────
def build_schema_tab(wb):
    ws = wb.create_sheet("Schema Analysis")

    # Header row — neutral background, bold text only
    for ci, name in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = BOLD
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[1].height = 20

    # Data rows
    with open(IN_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for ri, row in enumerate(reader, 2):
            for ci, col in enumerate(COLUMNS, 1):
                value = row.get(col, "")
                if col == "Recommendations":
                    value = format_recommendations(value)
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.alignment = WRAP
                cell.border = BORDER

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    return ws

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_schema_tab(wb)
    os.makedirs(os.path.dirname(OUT_XLS), exist_ok=True)
    wb.save(OUT_XLS)
    print(f"Saved: {OUT_XLS}")
    print(f"  Schema Analysis: 7 columns, sourced from {os.path.basename(IN_CSV)}")

if __name__ == "__main__":
    main()
